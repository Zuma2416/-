"""
Excel操作モジュール
立替経費精算書への明細追記機能
"""

import openpyxl
from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple


class ExpenseExcelHandler:
    """立替経費精算書のExcel操作クラス"""

    # テーブル構造定義
    HEADER_ROW = 10
    DATA_START_ROW = 11
    DATA_END_ROW = 26
    TOTAL_ROW = 27

    # 列定義（マージされたセルの開始列）
    COL_DATE = 1      # A列: 日付
    COL_PAYEE = 5     # E列: 支払先
    COL_CONTENT = 13  # M列: 支払内容
    COL_AMOUNT = 23   # W列: 金額

    def __init__(self, excel_path: str):
        """
        初期化

        Args:
            excel_path: Excelファイルのパス
        """
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excelファイルが見つかりません: {excel_path}")

        self.workbook = None
        self.worksheet = None

    def load(self):
        """Excelファイルを読み込む"""
        self.workbook = load_workbook(self.excel_path)
        self.worksheet = self.workbook.active

    def save(self, output_path: Optional[str] = None):
        """
        Excelファイルを保存

        Args:
            output_path: 出力先パス（Noneの場合は元ファイルを上書き）
        """
        if output_path:
            self.workbook.save(output_path)
        else:
            self.workbook.save(self.excel_path)

    def close(self):
        """Excelファイルを閉じる"""
        if self.workbook:
            self.workbook.close()

    def find_next_empty_row(self) -> Optional[int]:
        """
        次の空行を探す

        Returns:
            空行の行番号、または満杯の場合はNone
        """
        for row in range(self.DATA_START_ROW, self.DATA_END_ROW + 1):
            # 日付セルが空なら空行と判定
            date_cell = self.worksheet.cell(row, self.COL_DATE)
            if date_cell.value is None or str(date_cell.value).strip() == "":
                return row
        return None  # 満杯

    def add_expense_entry(
        self,
        date: str,
        payee: str,
        content: str,
        amount: float
    ) -> Tuple[bool, str]:
        """
        経費明細を追加

        Args:
            date: 日付（YYYY/MM/DD形式）
            payee: 支払先
            content: 支払内容
            amount: 金額

        Returns:
            (成功フラグ, メッセージ)
        """
        # 空行を探す
        empty_row = self.find_next_empty_row()
        if empty_row is None:
            return False, "明細行が満杯です（最大16件）"

        # データを書き込む
        try:
            self.worksheet.cell(empty_row, self.COL_DATE).value = date
            self.worksheet.cell(empty_row, self.COL_PAYEE).value = payee
            self.worksheet.cell(empty_row, self.COL_CONTENT).value = content
            self.worksheet.cell(empty_row, self.COL_AMOUNT).value = amount

            return True, f"{empty_row}行目に追加しました"
        except Exception as e:
            return False, f"書き込みエラー: {str(e)}"

    def get_existing_entries(self) -> List[Dict]:
        """
        既存の明細を取得

        Returns:
            明細のリスト
        """
        entries = []
        for row in range(self.DATA_START_ROW, self.DATA_END_ROW + 1):
            date_val = self.worksheet.cell(row, self.COL_DATE).value

            # 空行はスキップ
            if date_val is None or str(date_val).strip() == "":
                continue

            entry = {
                "row": row,
                "date": str(date_val),
                "payee": self.worksheet.cell(row, self.COL_PAYEE).value or "",
                "content": self.worksheet.cell(row, self.COL_CONTENT).value or "",
                "amount": self.worksheet.cell(row, self.COL_AMOUNT).value or 0
            }
            entries.append(entry)

        return entries

    def get_total_amount(self) -> float:
        """
        合計金額を取得

        Returns:
            合計金額
        """
        total_cell = self.worksheet.cell(self.TOTAL_ROW, self.COL_AMOUNT)
        total_value = total_cell.value

        # 数式の場合は計算結果を取得
        if isinstance(total_value, str) and total_value.startswith("="):
            # 数式を再計算（openpyxlでは自動計算されないため、手動計算）
            entries = self.get_existing_entries()
            return sum([float(e["amount"]) for e in entries if e["amount"]])

        return float(total_value) if total_value else 0.0

    def validate_data(self, date: str, payee: str, content: str, amount: float) -> Tuple[bool, str]:
        """
        データをバリデーション

        Args:
            date: 日付
            payee: 支払先
            content: 支払内容
            amount: 金額

        Returns:
            (有効フラグ, エラーメッセージ)
        """
        # 日付チェック
        try:
            datetime.strptime(date, "%Y/%m/%d")
        except ValueError:
            return False, "日付形式が不正です（YYYY/MM/DD形式で入力してください）"

        # 支払先チェック
        if not payee or len(payee) > 30:
            return False, "支払先は1〜30文字で入力してください"

        # 支払内容チェック
        if not content:
            return False, "支払内容を入力してください"

        # 金額チェック
        try:
            amount_float = float(amount)
            if amount_float <= 0:
                return False, "金額は正の数値で入力してください"
        except ValueError:
            return False, "金額が不正です"

        return True, ""


def test_excel_handler():
    """テスト関数"""
    handler = ExpenseExcelHandler("templates/立替経費精算書.xlsx")
    handler.load()

    print("=== 既存明細 ===")
    entries = handler.get_existing_entries()
    for entry in entries:
        print(f"行{entry['row']}: {entry['date']} | {entry['payee']} | {entry['content']} | {entry['amount']}")

    print(f"\n次の空行: {handler.find_next_empty_row()}")
    print(f"合計金額: {handler.get_total_amount()}")

    handler.close()


if __name__ == "__main__":
    test_excel_handler()
